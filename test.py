import openpyxl
from openpyxl.styles import PatternFill
import os
import re


def ekstrakcja(wyniki):
    with open(wyniki, 'r') as plik:
        dataline = re.compile('\w\s')
        plytka = []
        for linia in plik.readlines():
            if dataline.match(linia) is not None:
                plytka.append(linia.split())
    return plytka


def wypisz(plytki, ilosc, ws, pliki):
    for plytka, liczba in zip(plytki, range(ilosc)):
        for wiersz in range(0, 9):
            for kolumna in range(0, 13):
                if wiersz == 0:
                    if kolumna == 0:
                        a1 = ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1, value=pliki[liczba])
                        a1.fill = PatternFill(start_color='FFFF0000',
                                              end_color='FFFF0000',
                                              fill_type='solid')
                    else:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1, value=kolumna)
                else:
                    if kolumna == 0:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1, value=plytka[wiersz - 1][kolumna])
                    else:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1,
                                value=float(plytka[wiersz - 1][kolumna]))

def podaj_plytki(arkusze, opcja):
    for key in arkusze.keys():
        plytki = input(
            f'Podaj numery płytek z antygenem {key}, w których znajdują się surowice {opcja} (oddzielone spacją):\n')
        plytki = plytki.split()
        for i in range(len(plytki)):
            plytki[i] = int(plytki[i]) - 1
    return plytki


def podaj_studzienki(arkusze, plytki, opcja):
    slownik = {}
    for klucz in arkusze:
        slownik[klucz] = []
        for plytka in range(plytki[-1] + 1):
            if plytka in plytki:
                plus = input(f'Podaj początkową kolumnę i liczbę studzienek, w których znajdują się surowice {opcja}'
                             f' na płytce nr {plytka + 1} dla antygenu {klucz} (oddzielone spacją):\n')
                slownik[klucz].append(plus.split())
            else:
                slownik[klucz].append('k')
    return slownik


def srednie(plytki, studzienki, powtorzenia, ws, opcja):
    zakresy = []
    for liczba in plytki:
        for wiersz in range(0, 9):
            for kolumna in range(0, int(int(studzienki[liczba][1]) / (8 * powtorzenia)) + 1):
                if wiersz == 0:
                    if kolumna == 0:
                        a1 = ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 14 + int(studzienki[liczba][0]),
                                     value=f'średnia({opcja})')
                        a1.fill = PatternFill(start_color='FFFF0000',
                                              end_color='FFFF0000',
                                              fill_type='solid')
                else:
                    if kolumna == 0:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 14 + int(studzienki[liczba][0]),
                                value=f'{ws.cell(row=wiersz +1, column = 1).value}')
                    else:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 14 + int(studzienki[liczba][0]),
                                value=f'=AVERAGE('
                                      f'{ws.cell(row=wiersz + 1 + 10 * liczba, column=(kolumna - 1) * powtorzenia + int(studzienki[liczba][0]) + 1).coordinate}:'
                                      f'{ws.cell(row=wiersz + 1 + 10 * liczba, column=(kolumna - 1) * powtorzenia + int(studzienki[liczba][0]) + 1 + powtorzenia - 1).coordinate})')
        zakresy.append(f'{ws.cell(row=2 + 10 * liczba, column=14 + int(studzienki[liczba][0]) + 1).coordinate}:{ws.cell(row=9 + 10 * liczba, column=int(studzienki[liczba][1])/(8 * powtorzenia) + 14 + int(studzienki[liczba][0])).coordinate}')
    return zakresy

def wypisz_w_kolumnie(lista, it, ws, opcja):
    k = ws.cell(row=it, column=30, value=f'średnia({opcja})')
    k.fill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    for zakres in lista:
        for kol in range(len(tuple(ws[zakres])[0])):
            for wiersz in range(len(tuple(ws[zakres]))):
                ws.cell(row = it, column=31, value=tuple(ws[zakres])[wiersz][kol].value)
                it += 1
    return it

def cutoff(cutoffy, ws):
    temp = ''
    for zakres in cutoffy:
        if zakres == cutoffy[-1]:
            temp = temp + zakres
        else:
            temp = temp + zakres + ','
    c = ws.cell(row=1,
                column=31,
                value='cutoff')
    c.fill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    ws.cell(row=1,
            column=32,
            value=f'=AVERAGE({temp})+ 2 * (_xlfn.STDEVA({temp}))')


def czul_i_spec(it1, it2, it3, ws):
    kom_cutoff=ws.cell(row=1, column=32)
    for i in range(it1, it3):
        kom_sr = ws.cell(row=i, column=31)
        ws.cell(row=i, column=33, value=f'=IF({kom_sr.coordinate}>{kom_cutoff.coordinate},1,0)')
    zakres_czul=f'{ws.cell(row=it1, column=33).coordinate}:{ws.cell(row=it2-1, column=33).coordinate}'
    zakres_spec=f'{ws.cell(row=it2, column=33).coordinate}:{ws.cell(row=it3-1, column=33).coordinate}'
    czul_et = ws.cell(row=1, column=34, value='czułość')
    czul_et.fill = PatternFill(start_color='FFFF0000',
                                end_color='FFFF0000',
                                fill_type='solid')
    czul_wart = ws.cell(row=1, column=35, value=f'=(COUNTIF({zakres_czul},1)/COUNTA({zakres_czul}))*100')
    spec_et = ws.cell(row=1, column=36, value='specyficzność')
    spec_et.fill = PatternFill(start_color='FFFF0000',
                                end_color='FFFF0000',
                                fill_type='solid')
    spec_wart = ws.cell(row=1, column=37, value=f'=(COUNTIF({zakres_spec},0)/COUNTA({zakres_spec}))*100')



cwd = os.getcwd()
wb = openpyxl.Workbook()

arkusze = {}
ilosc_arkuszy = int(input('Podaj liczbę antygenów:\n'))
for i in range(1, ilosc_arkuszy + 1):
    nazwa_arkusza = input(f'Podaj nazwę antygenu nr {i}:\n')
    arkusze[nazwa_arkusza] = wb.create_sheet(nazwa_arkusza)
del wb['Sheet']

liczba_plytek = int(input('Podaj liczbę płytek na antygen:\n'))
pliki = {}
wyniki = {}
for key in arkusze.keys():
    pliki[key] = []
    wyniki[key] = []
    for plytka in range(1, liczba_plytek + 1):
        nazwa_pliku = input(
            f'Podaj nazwę pliku dla płyki nr {plytka} dla antygenu {key} (pamiętaj o rozszerzeniu .txt):\n')
        sciezka = os.path.join(cwd, nazwa_pliku)
        pliki[key].append(nazwa_pliku)
        wyniki[key].append(ekstrakcja(sciezka))

for key in arkusze.keys():
    ws = wb[key]
    wypisz(wyniki[key], liczba_plytek, ws, pliki[key])

powtorzenia = int(input('Podaj liczbę powtórzeń dla jednej surowicy:\n'))

plytki_pozytywne = podaj_plytki(arkusze, 'pozytywne')
plytki_negatywne = podaj_plytki(arkusze, 'negatywne')
plytki_cutoff = podaj_plytki(arkusze, 'do obliczenia cutoff')

studzienki_pozytywne = podaj_studzienki(arkusze, plytki_pozytywne, 'pozytywne')
studzienki_negatywne = podaj_studzienki(arkusze, plytki_negatywne, 'negatywne')
studzienki_cutoff = podaj_studzienki(arkusze, plytki_cutoff, 'do obliczenia cutoff')

for key in arkusze.keys():
    ws = wb[key]
    zakresy_plus = srednie(plytki_pozytywne, studzienki_pozytywne[key], powtorzenia, ws, '+')
    zakresy_minus = srednie(plytki_negatywne, studzienki_negatywne[key], powtorzenia, ws, '-')
    zakresy_co = srednie(plytki_cutoff, studzienki_cutoff[key], powtorzenia, ws, 'co')
    it = 2
    it_po_pl = wypisz_w_kolumnie(zakresy_plus, it, ws, '+')
    it_po_min = wypisz_w_kolumnie(zakresy_minus, it_po_pl, ws, '-')
    cutoff(zakresy_co, ws)
    czul_i_spec(it, it_po_pl, it_po_min, ws)


wb.save(r'C:\Users\Dudek\Desktop\test.xlsx')
wb.close()
