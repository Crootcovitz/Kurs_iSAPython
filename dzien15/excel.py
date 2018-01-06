import openpyxl
from pprint import pprint

# tak nie zadziała - trzeba tradycyjnie otworzyć i zamknąć
# with openpyxl.load_workbook(r'C:\Users\Dudek\Desktop\Kurs_Python\iSAPython4\dzien15\example.xlsx') as plik:
#     arkusze = plik.sheetnames
#     print(arkusze)

plik = openpyxl.load_workbook(r'C:\Users\Dudek\Desktop\Kurs_Python\iSAPython4\dzien15\example.xlsx')

arkusze = plik.sheetnames
print(arkusze)

aktywny_arkusz = plik.get_sheet_by_name('Owoce')
print(f'Aktywny arkusz: {aktywny_arkusz.title}.')

# komórki
komorka = aktywny_arkusz['A1']
print(komorka)
print(komorka.value)

# koordynaty
print(f'Adres komórki: {komorka.coordinate}.')
print(f'Kolumna komórki: {komorka.column}.')
print(f'Wiersz komórki: {komorka.row}.')

print(aktywny_arkusz.cell(row=2, column=2))

# zamiana liter kolumn na numer i vice versa
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(123))
print(column_index_from_string('ABC'))

# maksymalny rozmiar
print('Ostatnia kolumna:')
print(get_column_letter(aktywny_arkusz.max_column))
print('Ostatni wiersz:')
print(aktywny_arkusz.max_row)

# tuple w tuple, wewnętrzne to wiersze
pprint(aktywny_arkusz['A1':'C7'], indent=4)

for wiersz in aktywny_arkusz['A1':'C7']:
    for kom in wiersz:
        print(kom.value, end='\t\t')
    print()
print(aktywny_arkusz.cell(row=1, column=1).coordinate)
aktywny_arkusz['B9'] = 'Hello from Python!'
plik.save('example2.xlsx')
plik.close()
