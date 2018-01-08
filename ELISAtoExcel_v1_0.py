import openpyxl
from openpyxl.styles import PatternFill
import os
import re


def ekstrakcja(wyniki):
    with open(wyniki, 'r') as plik:
        dataline = re.compile('\w\s')
        plytka = []
        for linia in plik.readlines():
            if dataline.match(linia) is not None:
                plytka.append(linia.split())
    return plytka


def wypisz(plytki, ilosc, ws, pliki):
    for plytka, liczba in zip(plytki, range(ilosc)):
        for wiersz in range(0, 9):
            for kolumna in range(0, 13):
                if wiersz == 0:
                    if kolumna == 0:
                        a1 = ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1, value=pliki[liczba])
                        a1.fill = PatternFill(start_color='FFFF0000',
                                              end_color='FFFF0000',
                                              fill_type='solid')
                    else:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1, value=kolumna)
                else:
                    if kolumna == 0:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1, value=plytka[wiersz - 1][kolumna])
                    else:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 1,
                                value=float(plytka[wiersz - 1][kolumna]))

def podaj_plytki(arkusze, opcja):
    for key in arkusze.keys():
        check1 = True
        while check1:
            plytki = input(
                f'Podaj numery płytek z antygenem {key}, w których znajdują się surowice {opcja} (oddzielone spacją):\n')
            plytki = plytki.split()
            licznik = 0
            for i in range(len(plytki)):
                if plytki[i].isdigit():
                    licznik += 1
            if licznik == len(plytki):
                check1 = False
            else:
                print('Któraś z podanych wartości nie jest liczbą. Spróbuj ponownie.')
                continue
            for i in range(len(plytki)):
                plytki[i] = int(plytki[i]) - 1
    return plytki


def podaj_studzienki(arkusze, plytki, opcja):
    slownik = {}
    for klucz in arkusze:
        slownik[klucz] = []
        plytka = 0
        while plytka < plytki[-1] + 1:
            if plytka in plytki:
                plus = input(
                    f'Podaj początkową kolumnę i liczbę studzienek, w których znajdują się surowice {opcja}'
                    f' na płytce nr {plytka + 1} dla antygenu {klucz} (oddzielone spacją):\n')
                licznik = 0
                plus = plus.split()
                for i in range(len(plus)):
                    if plus[i].isdigit():
                        licznik += 1
                if licznik != len(plus):
                    print('Któraś z podanych wartości nie jest liczbą. Spróbuj ponownie.')
                    continue
                slownik[klucz].append(plus)
            else:
                slownik[klucz].append('k')
            plytka +=1
    return slownik


def srednie(plytki, studzienki, powtorzenia, ws, opcja):
    zakresy = []
    for liczba in plytki:
        for wiersz in range(0, 9):
            for kolumna in range(0, int(int(studzienki[liczba][1]) / (8 * powtorzenia)) + 1):
                if wiersz == 0:
                    if kolumna == 0:
                        a1 = ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 14 + int(studzienki[liczba][0]),
                                     value=f'średnia({opcja})')
                        a1.fill = PatternFill(start_color='FFFF0000',
                                              end_color='FFFF0000',
                                              fill_type='solid')
                else:
                    if kolumna == 0:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 14 + int(studzienki[liczba][0]),
                                value=f'{ws.cell(row=wiersz +1, column = 1).value}')
                    else:
                        ws.cell(row=wiersz + 1 + 10 * liczba, column=kolumna + 14 + int(studzienki[liczba][0]),
                                value=f'=AVERAGE('
                                      f'{ws.cell(row=wiersz + 1 + 10 * liczba, column=(kolumna - 1) * powtorzenia + int(studzienki[liczba][0]) + 1).coordinate}:'
                                      f'{ws.cell(row=wiersz + 1 + 10 * liczba, column=(kolumna - 1) * powtorzenia + int(studzienki[liczba][0]) + 1 + powtorzenia - 1).coordinate})')
        zakresy.append(f'{ws.cell(row=2 + 10 * liczba, column=14 + int(studzienki[liczba][0]) + 1).coordinate}:{ws.cell(row=9 + 10 * liczba, column=int(studzienki[liczba][1])/(8 * powtorzenia) + 14 + int(studzienki[liczba][0])).coordinate}')
    return zakresy

def wypisz_w_kolumnie(lista, it, ws, opcja):
    k = ws.cell(row=it, column=30, value=f'średnia({opcja})')
    k.fill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    for zakres in lista:
        for kol in range(len(tuple(ws[zakres])[0])):
            for wiersz in range(len(tuple(ws[zakres]))):
                ws.cell(row = it, column=31, value=tuple(ws[zakres])[wiersz][kol].value)
                it += 1
    return it

def cutoff(cutoffy, ws):
    temp = ''
    for zakres in cutoffy:
        if zakres == cutoffy[-1]:
            temp = temp + zakres
        else:
            temp = temp + zakres + ','
    c = ws.cell(row=1,
                column=31,
                value='cutoff')
    c.fill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    ws.cell(row=1,
            column=32,
            value=f'=AVERAGE({temp})+ 2 * (_xlfn.STDEVA({temp}))')


def czul_i_spec(it1, it2, it3, ws):
    kom_cutoff=ws.cell(row=1, column=32)
    for i in range(it1, it3):
        kom_sr = ws.cell(row=i, column=31)
        ws.cell(row=i, column=33, value=f'=IF({kom_sr.coordinate}>{kom_cutoff.coordinate},1,0)')
    zakres_czul=f'{ws.cell(row=it1, column=33).coordinate}:{ws.cell(row=it2-1, column=33).coordinate}'
    zakres_spec=f'{ws.cell(row=it2, column=33).coordinate}:{ws.cell(row=it3-1, column=33).coordinate}'
    czul_et = ws.cell(row=1, column=34, value='czułość')
    czul_et.fill = PatternFill(start_color='FFFF0000',
                                end_color='FFFF0000',
                                fill_type='solid')
    czul_wart = ws.cell(row=1, column=35, value=f'=(COUNTIF({zakres_czul},1)/COUNTA({zakres_czul}))*100')
    spec_et = ws.cell(row=1, column=36, value='specyficzność')
    spec_et.fill = PatternFill(start_color='FFFF0000',
                                end_color='FFFF0000',
                                fill_type='solid')
    spec_wart = ws.cell(row=1, column=37, value=f'=(COUNTIF({zakres_spec},0)/COUNTA({zakres_spec}))*100')

print('Program ELISAtoExcel_v1_0 napisany przez: Mateusz Dutkiewicz\n\n'
      'Program przenosi wyniki pomiarów absorbancji w teście ELISA do skoroszytu Excel oraz oblicza czułość i specyficzność.\n'
      'Wszystkie studzienki na płytce muszą być wypełnione.\n'
      'Przed podjęciem dalszych działań przenieś wszystkie pliki z wynikami do folderu, w którym znajduje się program.\n\n')

cwd = os.getcwd()
wb = openpyxl.Workbook()
checkpoint1 = True
while checkpoint1:
    arkusze = {}
    ilosc_arkuszy = input('Podaj liczbę antygenów:\n')
    if ilosc_arkuszy.isdigit():
        ilosc_arkuszy = int(ilosc_arkuszy)
    else:
        print('Podany znak nie jest liczbą. Spróbuj ponownie.')
        continue
    for i in range(1, ilosc_arkuszy + 1):
        nazwa_arkusza = input(f'Podaj nazwę antygenu nr {i}:\n')
        arkusze[nazwa_arkusza] = wb.create_sheet(nazwa_arkusza)
    print(f'Utworzono arkusze dla {ilosc_arkuszy} antygenów o nazwach:')
    for key in arkusze.keys():
        print(key)
    check_pyt1 = True
    while check_pyt1:
        pyt1 = input('Kontynuować (y), czy powtórzyć ten krok (n)?\n')
        if pyt1 == 'y':
            check_pyt1 = False
            checkpoint1 = False
        elif pyt1 == 'n':
            check_pyt1 = False
        else:
            print('Nie zrozumiałem. Spróbuj ponownie.')

del wb['Sheet']

checkpoint2 = True
while checkpoint2:
    liczba_plytek = input('Podaj liczbę płytek na antygen:\n')
    if liczba_plytek.isdigit():
        liczba_plytek = int(liczba_plytek)
    else:
        print('Podany znak nie jest liczbą. Spróbuj ponownie.')
        continue
    pliki = {}
    wyniki = {}
    for key in arkusze.keys():
        pliki[key] = []
        wyniki[key] = []
        plytka = 1
        while plytka < liczba_plytek + 1:
            nazwa_pliku = input(
                f'Podaj nazwę pliku dla płyki nr {plytka} dla antygenu {key} (pamiętaj o rozszerzeniu .txt):\n')
            sciezka = os.path.join(cwd, nazwa_pliku)
            pliki[key].append(nazwa_pliku)
            try:
                wyniki[key].append(ekstrakcja(sciezka))
                plytka += 1
            except FileNotFoundError:
                print('Pliku nie znaleziono. Spróbuj ponownie.')
                continue
    print('Wczytano wszystkie pliki.')
    check_pyt2 = True
    while check_pyt2:
        pyt2 = input('Kontynuować (y), czy powtórzyć ten krok (n)?\n')
        if pyt2 == 'y':
            check_pyt2 = False
            checkpoint2 = False
        elif pyt2 == 'n':
            check_pyt2 = False
        else:
            print('Nie zrozumiałem. Spróbuj ponownie.')


for key in arkusze.keys():
    ws = wb[key]
    wypisz(wyniki[key], liczba_plytek, ws, pliki[key])
print('Pomyślnie wpisano dane do arkusza.')

checkpoint3 = True
while checkpoint3:
    powtorzenia = input('Podaj liczbę powtórzeń dla jednej surowicy:\n')
    if powtorzenia.isdigit():
        powtorzenia = int(powtorzenia)
    else:
        print('Podany znak nie jest liczbą. Spróbuj ponownie.')
        continue
    plytki_pozytywne = podaj_plytki(arkusze, 'pozytywne')
    plytki_negatywne = podaj_plytki(arkusze, 'negatywne')
    plytki_cutoff = podaj_plytki(arkusze, 'do obliczenia cutoff')

    studzienki_pozytywne = podaj_studzienki(arkusze, plytki_pozytywne, 'pozytywne')
    studzienki_negatywne = podaj_studzienki(arkusze, plytki_negatywne, 'negatywne')
    studzienki_cutoff = podaj_studzienki(arkusze, plytki_cutoff, 'do obliczenia cutoff')
    check_pyt3 = True
    while check_pyt3:
        pyt3 = input('Kontynuować (y), czy powtórzyć ten krok (n)?\n')
        if pyt3 == 'y':
            check_pyt3 = False
            checkpoint3 = False
        elif pyt3 == 'n':
            check_pyt3 = False
        else:
            print('Nie zrozumiałem. Spróbuj ponownie.')

for key in arkusze.keys():
    ws = wb[key]
    zakresy_plus = srednie(plytki_pozytywne, studzienki_pozytywne[key], powtorzenia, ws, '+')
    zakresy_minus = srednie(plytki_negatywne, studzienki_negatywne[key], powtorzenia, ws, '-')
    zakresy_co = srednie(plytki_cutoff, studzienki_cutoff[key], powtorzenia, ws, 'co')
    it = 2
    it_po_pl = wypisz_w_kolumnie(zakresy_plus, it, ws, '+')
    it_po_min = wypisz_w_kolumnie(zakresy_minus, it_po_pl, ws, '-')
    cutoff(zakresy_co, ws)
    czul_i_spec(it, it_po_pl, it_po_min, ws)
print('Pomyślnie przetworzono dane.')

checkpoint4 = True
while checkpoint4:
    nazwa_skoroszytu = input('Podaj nazwę pliku wyjściowego (pamiętaj o rozszerzeniu .xlsx):\n')
    sciezka_zapisu = os.path.join(cwd, nazwa_skoroszytu)
    try:
        wb.save(sciezka_zapisu)
        wb.close()
    except PermissionError:
        print('Odmowa dostępu - plik o podanej nazwie jest w tej chwili otwarty. Zamknij go lub użyj innej nazwy.')
        continue
    checkpoint4 = False
koniec = input('Pomyślnie zapisano plik. Naciśnij ENTER, aby zakończyć program.\n')
